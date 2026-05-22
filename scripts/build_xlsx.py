"""progress.json → vocabulary.xlsx 변환.

이미지 인덱스 순서대로 단어/뜻을 정렬하여 단일 시트 .xlsx 생성.
"""
from __future__ import annotations
import argparse
import json
import sys
from pathlib import Path

from openpyxl import Workbook


def build(progress_path: Path, output_path: Path) -> int:
    data = json.loads(progress_path.read_text(encoding="utf-8"))
    images = data.get("images", {})

    wb = Workbook()
    ws = wb.active
    ws.title = "Vocabulary"
    ws.cell(row=1, column=1, value="단어")
    ws.cell(row=1, column=2, value="뜻")

    row = 2
    total = 0
    for idx in sorted(images.keys()):
        for pair in images[idx]:
            ws.cell(row=row, column=1, value=pair["word"])
            ws.cell(row=row, column=2, value=pair["meaning"])
            row += 1
            total += 1

    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 50

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    print(f"작성 완료: {output_path} ({total}개 단어)")
    return total


def main(argv=None) -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True)
    parser.add_argument("--output", required=True)
    args = parser.parse_args(argv)
    build(Path(args.input), Path(args.output))
    return 0


if __name__ == "__main__":
    sys.exit(main())
